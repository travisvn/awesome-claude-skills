from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

NAVY = "1E2761"; TINT = "F4F6F8"; RULE = "C8CDD3"; ADA = "E8F1F5"
thin = Side(style="thin", color=RULE)
box = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()
ws = wb.active
ws.title = "Material Requirement"


def cell(r, c, v, *, bold=False, size=10, fill=None, align="left", wrap=False,
         color="1F2937", fmt=None, border=True):
    x = ws.cell(row=r, column=c, value=v)
    x.font = Font(name="Arial", size=size, bold=bold, color=color)
    if fill:
        x.fill = PatternFill("solid", fgColor=fill)
    x.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if border:
        x.border = box
    if fmt:
        x.number_format = fmt
    return x


# ---------------------------------------------------------------- title block
cell(1, 1, "MATERIAL REQUIREMENT — AGL INSTALLATION", bold=True, size=14,
     color=NAVY, border=False)
cell(2, 1, "Rectification of TWY E between E4 & E6 (ZIA) — AGL works at three milling "
           "locations", size=10, color="5F6368", border=False)
meta = [
    ("Document No.", "AUH-SK-AGL-TWYE-001 — Rev P08 (final scope of work)"),
    ("Scope basis", "Rev P08 final scope of work issued 30.07.2026 — 19 No. affected fittings; "
                    "13 No. new side-entry shallow bases; approx. 420 m of saw cut"),
    ("Sourcing", "Two lines fall short of ADB SAFEGATE stock. The project team confirms no "
                 "ready stock; both are held in ADA inventory and are to be drawn from there."),
    ("Prepared by", "Mohammed Faheem — AGL Team Leader, ADB SAFEGATE"),
    ("Approved by", "Ragesh Menon — AGL Manager, ADB SAFEGATE"),
    ("Date", "31.07.2026"),
]
for i, (k, v) in enumerate(meta, start=4):
    cell(i, 1, k, bold=True, size=10, color="AD841F", border=False)
    cell(i, 2, v, size=10, border=False)

# ---------------------------------------------------------------- table
HDR = ["Sl No", "Material", "ADA item code", "Unit", "Required Qty",
       "Available with ADB SAFEGATE", "Qty requested from ADA inventory",
       "ADA inventory — current stock", "ADA inventory — balance after withdrawal",
       "Basis / check against the Rev P08 plan"]
HR = 11
for c, h in enumerate(HDR, start=1):
    cell(HR, c, h, bold=True, size=10, fill=NAVY, color="FFFFFF",
         align="center" if c != 10 else "left", wrap=True)

# material, ADA code, unit, required, available with ADB SG, ADA stock, basis
ROWS = [
    ("Nitoseal", None, "ltr", 20, 20, None,
     "Saw cut sealant. Against approx. 420 m of cut this is approx. 0.05 l/m — cannot be "
     "firmed until the saw cut detail fixes cut width and depth (hold point H5)."),
    ("Nito mortar", None, "ltr", 15, 15, None,
     "Grout / bedding to the new bases — 13 No. at approx. 1.15 l each. Depends on the core "
     "diameter, which is pending the same detail."),
    ("Backer rod", None, "Roll", 8, 8, None,
     "Approx. 400 m at 50 m/roll against approx. 420 m of saw cut — confirm the roll length; "
     "a 9th roll may be needed."),
    ("Shallow base 8 inch", "10016966", "No", 11, 4, 10,
     "11 No. 8\" side-entry shallow bases at LOC-01. 7 No. short of ADB SAFEGATE stock and "
     "drawn from ADA inventory, which holds 10 No. — 3 No. remain after withdrawal."),
    ("Earth cable", None, "Mtr", 900, 900, None,
     "As advised. No earthing quantity is derived in the Rev P08 deck."),
    ("Secondary connectors (plug & receptacle)", None, "Pair", 38, 38, None,
     "2 pair per fitting x 19 No. = 38 No."),
    ("Secondary cable (2c x 4 sq.mm)", "10020584", "mtr", 950, 600, 2000,
     "1 No. 2-core 4 sq.mm cable to each fitting, SBC and TCC alike — 7 No. SBC and 12 No. "
     "TCC = 19 No. runs. 350 m short of ADB SAFEGATE stock and drawn from ADA inventory, "
     "which holds 1 RoL @ 2000 m — 1650 m remain on the roll after withdrawal."),
    ("Masking tape", None, "Box", 50, 50, None,
     "As advised, for general protection. The signboard masking is by matt black vinyl "
     "sticker — see note 7; there is no material line for it in this table."),
    ("Shallow base 12 inch", None, "No", 2, 2, None,
     "The plan needs 13 No. new bases: 11 No. 8\" plus 2 No. 12\" (LOC-02 TCCECH-03/008 and "
     "LOC-03 TCCECH-03/003)."),
    ("Dummy plate (8\" / 12\")", None, "No", 16, 16, None,
     "13 No. open bases plus the 3 No. LOC-01 fittings kept under a plate during milling "
     "(Phase 1, R3)."),
]
r = HR + 1
for i, (mat, code, unit, req, avail, ada, basis) in enumerate(ROWS, start=1):
    drawn = ada is not None
    fill = ADA if drawn else (TINT if i % 2 == 0 else None)
    cell(r, 1, i, align="center", fill=fill)
    cell(r, 2, mat, fill=fill, wrap=True, bold=drawn)
    cell(r, 3, code, align="center", fill=fill)
    cell(r, 4, unit, align="center", fill=fill)
    cell(r, 5, req, align="center", fill=fill, fmt="#,##0")
    cell(r, 6, avail, align="center", fill=fill, fmt="#,##0")
    cell(r, 7, f"=MAX(0,E{r}-F{r})", align="center", fill=fill, fmt="#,##0", bold=True)
    cell(r, 8, ada, align="center", fill=fill, fmt="#,##0")
    cell(r, 9, f"=IF(H{r}=\"\",\"\",H{r}-G{r})", align="center", fill=fill, fmt="#,##0",
         bold=True)
    cell(r, 10, basis, size=9, wrap=True, color="5F6368", fill=fill)
    r += 1

TOT = r
for c in range(1, 11):
    cell(TOT, c, None, fill=TINT)
cell(TOT, 2, "Lines drawn from ADA inventory", bold=True, fill=TINT)
cell(TOT, 7, f"=COUNTIF(G{HR + 1}:G{TOT - 1},\">0\")", bold=True, align="center", fill=TINT)
cell(TOT, 10, "Every other line is covered by ADB SAFEGATE stock.", size=9,
     color="5F6368", wrap=True, fill=TINT)

# ---------------------------------------------------------- ADA inventory detail
d = TOT + 2
cell(d, 1, "ADA INVENTORY — ITEMS TO BE WITHDRAWN", bold=True, size=11, color=NAVY,
     border=False)
for c, h in enumerate(["Ref", "Item code", "Description as held in ADA inventory", "Unit",
                       "Stock", "Withdrawal"], start=1):
    cell(d + 1, c, h, bold=True, size=10, fill=NAVY, color="FFFFFF",
         align="center" if c != 3 else "left", wrap=True)
DROWS = [
    ("D02", "10016966",
     "BASE, MOUNTING; TYP: HIGH PRESSURE INJECTION SHALLOW, DIA 8IN; MFR: ADB SAFEGATE; "
     "P/N MSBC122V0003; M10; HS 85309000; with earthing connection, fixing screw set, one "
     "side cable entry (pole qty 2)", "Nos.", 10, "7 No."),
    ("D03", "10020584",
     "CABLE, ELECTRICAL; CNDCTR DIA 4MM2; CNDCTR QTY 2C; MFR: EUPEN; 1 RoL / 2000 m",
     "RoL", 1, "350 m off the roll"),
]
for n_i, row in enumerate(DROWS):
    rr = d + 2 + n_i
    for c, v in enumerate(row, start=1):
        cell(rr, c, v, size=9, wrap=(c == 3),
             align="center" if c in (1, 4, 5) else "left", fill=ADA)
    ws.row_dimensions[rr].height = 46
ws.merge_cells(start_row=d + 1, start_column=6, end_row=d + 1, end_column=10)
for n_i in range(len(DROWS)):
    ws.merge_cells(start_row=d + 2 + n_i, start_column=6, end_row=d + 2 + n_i, end_column=10)

# ---------------------------------------------------------------- notes
n = d + 2 + len(DROWS) + 1
cell(n, 1, "NOTES", bold=True, size=11, color=NAVY, border=False)
NOTES = [
    "1.  Two lines fall short of ADB SAFEGATE stock — 7 No. shallow base 8 inch and 350 m of "
    "secondary cable. The project team confirms no ready stock; both are held in ADA inventory "
    "and are to be drawn from there under items 10016966 and 10020584.",
    "2.  ADA inventory balance after withdrawal: 3 No. shallow base 8 inch, and 1650 m "
    "remaining on the 2000 m cable roll. The roll is issued whole, so the balance stays on it.",
    "3.  Confirm that ADA item 10016966 is the side-entry type the plan calls for. The stores "
    "description reads 'one side cable entry', which is consistent, but the part number should "
    "be checked against the base required before withdrawal.",
    "4.  Confirm the rating of ADA item 10020584 against the ADA specification for the series "
    "secondary circuit. The stores description reads 2750.1 kV, which appears to be a "
    "data-entry artefact rather than the cable rating.",
    "5.  The plan needs 13 No. new side-entry shallow bases in total — 11 No. 8\" at LOC-01, "
    "1 No. 12\" at LOC-02 (TCCECH-03/008) and 1 No. 12\" at LOC-03 (TCCECH-03/003).",
    "6.  Sealant, mortar and backer rod quantities depend on the saw cut width and depth and on "
    "the core diameter. Both are set by the saw cut & side-entry shallow base detail drawing, "
    "which had not been issued at Rev P08 (hold point H5). Re-check these three lines when it "
    "is issued.",
    "7.  The direction signboards leading to E4 and E6 are masked with matt black vinyl sticker "
    "under Phase 1 (R4) and unmasked in Phase 3 before the functionality check. There is no "
    "material line for the vinyl in this table — add one once the sign faces are measured.",
    "8.  One 2-core 4 sq.mm secondary cable serves each fitting, SBC and TCC alike — 7 No. SBC "
    "and 12 No. TCC. Cable is ordered against the full manhole-to-light runs (19 No.), not the "
    "saw cut length — the cut measures approx. 420 m across the three locations (approx. 300 m "
    "LOC-01, 24 m LOC-02, 95 m LOC-03).",
    "9.  Saw cut is an interim arrangement agreed between the AGL team, the civil team and "
    "ADA AGL. Permanent duct provision follows under the South Rehabilitation works, and is "
    "not covered by this requirement.",
    "10.  Testing and commissioning is carried out at circuit level and covers the affected "
    "fittings together with the fittings on the same circuits that fall outside these works — "
    "the series circuits are proved end to end before handover.",
    "11.  Qty requested from ADA inventory and the balance after withdrawal are formulas. Edit "
    "only Required Qty, Available with ADB SAFEGATE, and ADA inventory current stock.",
]
for i, t in enumerate(NOTES):
    c = cell(n + 1 + i, 1, t, size=9, color="5F6368", wrap=True, border=False)
    ws.merge_cells(start_row=n + 1 + i, start_column=1, end_row=n + 1 + i, end_column=10)
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[n + 1 + i].height = 26

for col, w in zip("ABCDEFGHIJ", (6, 30, 12, 7, 11, 15, 15, 15, 17, 52)):
    ws.column_dimensions[col].width = w
for rr in range(HR, TOT + 1):
    ws.row_dimensions[rr].height = 40
ws.row_dimensions[HR].height = 54
ws.freeze_panes = "A12"
ws.print_title_rows = "11:11"
ws.page_setup.orientation = "landscape"
ws.page_setup.fitToWidth = 1
ws.sheet_properties.pageSetUpPr.fitToPage = True

wb.save("Material Requirement - TWY E AGL Installation - Rev P08.xlsx")
print("saved")
